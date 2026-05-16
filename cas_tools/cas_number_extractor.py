from __future__ import annotations

import csv
import io
import re
import zipfile
from collections import defaultdict
from dataclasses import dataclass
from pathlib import Path
from typing import List, Optional, Sequence, Tuple
from xml.etree import ElementTree as ET


CAS_PATTERN = re.compile(r"(?<!\d)(\d{2,7}-\d{2}-\d)(?!\d)")
TEXT_ENCODINGS: Tuple[str, ...] = (
    "utf-8-sig",
    "utf-8",
    "utf-16",
    "utf-16-le",
    "utf-16-be",
    "cp1252",
    "latin-1",
)
TEXT_EXTENSIONS = {
    ".txt",
    ".md",
    ".rst",
    ".log",
    ".text",
    ".json",
    ".yaml",
    ".yml",
    ".xml",
    ".html",
    ".htm",
    ".csv",
    ".tsv",
}
WORD_EXTENSIONS = {".docx", ".docm"}
EXCEL_OPENXML_EXTENSIONS = {".xlsx", ".xlsm"}
EXCEL_LEGACY_EXTENSIONS = {".xls"}
PDF_EXTENSIONS = {".pdf"}
SUPPORTED_EXTENSIONS = TEXT_EXTENSIONS | WORD_EXTENSIONS | EXCEL_OPENXML_EXTENSIONS | EXCEL_LEGACY_EXTENSIONS | PDF_EXTENSIONS
OFFICE_TEMP_PREFIX = "~$"
MAX_GENERIC_TEXT_SNIFF_BYTES = 4096
MAX_CONTEXT_LENGTH = 180
_WORD_NS = {"w": "http://schemas.openxmlformats.org/wordprocessingml/2006/main"}
_XLSX_NS = {
    "main": "http://schemas.openxmlformats.org/spreadsheetml/2006/main",
    "rel": "http://schemas.openxmlformats.org/officeDocument/2006/relationships",
}
_PKG_REL_NS = "http://schemas.openxmlformats.org/package/2006/relationships"


@dataclass(frozen=True)
class CASMatch:
    source_file: str
    relative_path: str
    file_type: str
    location: str
    cas_number: str
    context: str

    def to_row(self) -> dict[str, str]:
        return {
            "source_file": self.source_file,
            "relative_path": self.relative_path,
            "file_type": self.file_type,
            "location": self.location,
            "cas_number": self.cas_number,
            "context": self.context,
        }


def is_valid_cas_number(value: str) -> bool:
    text = str(value or "").strip()
    if not CAS_PATTERN.fullmatch(text):
        return False
    digits = text.replace("-", "")
    if len(digits) < 5:
        return False
    checksum = int(digits[-1])
    total = 0
    for multiplier, digit in enumerate(reversed(digits[:-1]), start=1):
        total += multiplier * int(digit)
    return total % 10 == checksum


def find_cas_numbers_in_text(text: str) -> List[str]:
    ordered: List[str] = []
    seen: set[str] = set()
    for match in CAS_PATTERN.finditer(text or ""):
        cas_number = match.group(1)
        if cas_number in seen or not is_valid_cas_number(cas_number):
            continue
        seen.add(cas_number)
        ordered.append(cas_number)
    return ordered


def discover_candidate_files(folder_path: str | Path, excluded_paths: Optional[Sequence[str | Path]] = None) -> List[Path]:
    folder = Path(folder_path)
    excluded = {Path(item).resolve() for item in (excluded_paths or [])}
    candidates: List[Path] = []
    for path in sorted(folder.rglob("*")):
        if not path.is_file():
            continue
        if path.name.startswith(OFFICE_TEMP_PREFIX):
            continue
        try:
            resolved = path.resolve()
        except Exception:
            resolved = path
        if resolved in excluded:
            continue
        if _is_candidate_file(path):
            candidates.append(path)
    return candidates


def extract_cas_matches_from_file(path: str | Path, base_folder: str | Path | None = None) -> Tuple[List[CASMatch], List[str]]:
    file_path = Path(path)
    relative_path = _relative_path(file_path, base_folder)
    suffix = file_path.suffix.lower()
    try:
        if suffix in PDF_EXTENSIONS:
            return _extract_from_pdf(file_path, relative_path), []
        if suffix in {".csv", ".tsv"}:
            return _extract_from_delimited_text(file_path, relative_path), []
        if suffix in WORD_EXTENSIONS:
            return _extract_from_docx(file_path, relative_path), []
        if suffix in EXCEL_OPENXML_EXTENSIONS:
            return _extract_from_xlsx(file_path, relative_path), []
        if suffix in EXCEL_LEGACY_EXTENSIONS:
            return _extract_from_xls(file_path, relative_path)
        if suffix in TEXT_EXTENSIONS or _looks_like_text_file(file_path):
            return _extract_from_text_file(file_path, relative_path), []
        return [], [f"Skipped unsupported file: {file_path}"]
    except Exception as exc:
        return [], [f"Failed to process {file_path}: {exc}"]


def write_matches_to_csv(matches: Sequence[CASMatch], output_path: str | Path) -> None:
    path = Path(output_path)
    path.parent.mkdir(parents=True, exist_ok=True)
    existing_cas_numbers = _read_existing_cas_numbers(path)
    ordered_new_cas_numbers: List[str] = []
    seen_in_batch: set[str] = set()

    for match in matches:
        cas_number = match.cas_number
        if cas_number in existing_cas_numbers or cas_number in seen_in_batch:
            continue
        seen_in_batch.add(cas_number)
        ordered_new_cas_numbers.append(cas_number)

    file_exists = path.exists()
    needs_header = not file_exists or path.stat().st_size == 0
    mode = "a" if file_exists else "w"
    with path.open(mode, newline="", encoding="utf-8") as handle:
        writer = csv.writer(handle)
        if needs_header:
            writer.writerow(["cas_number"])
        for cas_number in ordered_new_cas_numbers:
            writer.writerow([cas_number])


def _read_existing_cas_numbers(path: Path) -> set[str]:
    if not path.exists() or path.stat().st_size == 0:
        return set()

    existing: set[str] = set()
    with path.open("r", newline="", encoding="utf-8") as handle:
        reader = csv.reader(handle)
        header_checked = False
        cas_column_index = 0
        for row in reader:
            if not row:
                continue
            if not header_checked:
                header_checked = True
                normalized_header = [cell.strip() for cell in row]
                if "cas_number" in normalized_header:
                    cas_column_index = normalized_header.index("cas_number")
                    continue
                if len(normalized_header) == 1 and normalized_header[0].lower() == "cas_number":
                    continue
            if cas_column_index >= len(row):
                continue
            cas_number = str(row[cas_column_index]).strip()
            if is_valid_cas_number(cas_number):
                existing.add(cas_number)
    return existing


def write_matches_to_markdown(
    matches: Sequence[CASMatch],
    output_path: str | Path,
    *,
    source_folder: str = "",
    processed_files: Optional[Sequence[str]] = None,
    warnings: Optional[Sequence[str]] = None,
) -> None:
    path = Path(output_path)
    path.parent.mkdir(parents=True, exist_ok=True)

    sorted_matches = sorted(
        matches,
        key=lambda item: (item.relative_path.lower(), item.location.lower(), item.cas_number),
    )
    unique_cas = []
    seen_cas: set[str] = set()
    for match in sorted_matches:
        if match.cas_number in seen_cas:
            continue
        seen_cas.add(match.cas_number)
        unique_cas.append(match.cas_number)

    grouped_matches: dict[str, List[CASMatch]] = defaultdict(list)
    for match in sorted_matches:
        grouped_matches[match.relative_path].append(match)

    with path.open("w", encoding="utf-8") as handle:
        report_name = source_folder or path.stem
        handle.write(f"# CAS Number Extraction Report ({report_name})\n\n")
        if processed_files is not None:
            handle.write(f"Total files processed: {len(processed_files)}\n\n")
        handle.write(f"Total CAS matches: {len(sorted_matches)}\n\n")
        handle.write(f"Unique CAS numbers: {len(unique_cas)}\n\n")

        if unique_cas:
            handle.write("## Unique CAS Numbers\n\n")
            for cas_number in unique_cas:
                handle.write(f"- {cas_number}\n")
            handle.write("\n")

        warning_list = [warning for warning in (warnings or []) if str(warning).strip()]
        if warning_list:
            handle.write("## Warnings\n\n")
            for warning in warning_list:
                handle.write(f"- {_escape_markdown_text(str(warning))}\n")
            handle.write("\n")

        handle.write("## File Details\n\n")
        if not grouped_matches:
            handle.write("No CAS numbers were found.\n")
            return

        for relative_path, file_matches in grouped_matches.items():
            first_match = file_matches[0]
            handle.write(f"### {_escape_markdown_text(relative_path)}\n\n")
            handle.write(f"- File type: {first_match.file_type}\n")
            handle.write(f"- Matches: {len(file_matches)}\n\n")
            for match in file_matches:
                context = _escape_markdown_text(match.context)
                handle.write(f"- {match.location}: {match.cas_number} | {context}\n")
            handle.write("\n")


def _escape_markdown_text(text: str) -> str:
    return str(text or "").replace("\r", " ").replace("\n", " ").strip()


def _is_candidate_file(path: Path) -> bool:
    suffix = path.suffix.lower()
    if suffix in SUPPORTED_EXTENSIONS:
        return True
    return _looks_like_text_file(path)


def _looks_like_text_file(path: Path) -> bool:
    try:
        with path.open("rb") as handle:
            sample = handle.read(MAX_GENERIC_TEXT_SNIFF_BYTES)
    except Exception:
        return False
    if not sample:
        return True
    if b"\x00" in sample:
        return False
    for encoding in TEXT_ENCODINGS:
        try:
            sample.decode(encoding)
            return True
        except UnicodeDecodeError:
            continue
    return False


def _relative_path(path: Path, base_folder: str | Path | None) -> str:
    if base_folder is None:
        return path.name
    try:
        return str(path.resolve().relative_to(Path(base_folder).resolve()))
    except Exception:
        return str(path)


def _read_text_bytes(path: Path) -> str:
    raw_bytes = path.read_bytes()
    for encoding in TEXT_ENCODINGS:
        try:
            return raw_bytes.decode(encoding)
        except UnicodeDecodeError:
            continue
    return raw_bytes.decode("utf-8", errors="replace")


def _compact_context(text: str, start: Optional[int] = None, end: Optional[int] = None) -> str:
    normalized = re.sub(r"\s+", " ", text or "").strip()
    if len(normalized) <= MAX_CONTEXT_LENGTH:
        return normalized
    if start is None or end is None:
        return normalized[: MAX_CONTEXT_LENGTH - 3] + "..."
    center = max((start + end) // 2, 0)
    left = max(center - (MAX_CONTEXT_LENGTH // 2), 0)
    right = min(left + MAX_CONTEXT_LENGTH, len(normalized))
    snippet = normalized[left:right]
    if left > 0:
        snippet = "..." + snippet[3:]
    if right < len(normalized):
        snippet = snippet[:-3] + "..."
    return snippet


def _matches_from_text(text: str, *, source_file: str, relative_path: str, file_type: str, location: str) -> List[CASMatch]:
    matches: List[CASMatch] = []
    for match in CAS_PATTERN.finditer(text or ""):
        cas_number = match.group(1)
        if not is_valid_cas_number(cas_number):
            continue
        matches.append(
            CASMatch(
                source_file=source_file,
                relative_path=relative_path,
                file_type=file_type,
                location=location,
                cas_number=cas_number,
                context=_compact_context(text, match.start(), match.end()),
            )
        )
    return matches


def _extract_from_text_file(path: Path, relative_path: str) -> List[CASMatch]:
    text = _read_text_bytes(path)
    matches: List[CASMatch] = []
    for line_number, line in enumerate(text.splitlines(), start=1):
        matches.extend(
            _matches_from_text(
                line,
                source_file=path.name,
                relative_path=relative_path,
                file_type="text",
                location=f"line {line_number}",
            )
        )
    return matches


def _extract_from_delimited_text(path: Path, relative_path: str) -> List[CASMatch]:
    text = _read_text_bytes(path)
    delimiter = "\t" if path.suffix.lower() == ".tsv" else ","
    reader = csv.reader(io.StringIO(text), delimiter=delimiter)
    matches: List[CASMatch] = []
    for row_index, row in enumerate(reader, start=1):
        for column_index, cell in enumerate(row, start=1):
            matches.extend(
                _matches_from_text(
                    str(cell),
                    source_file=path.name,
                    relative_path=relative_path,
                    file_type=path.suffix.lower().lstrip("."),
                    location=f"row {row_index}, col {column_index}",
                )
            )
    return matches


def _extract_from_pdf(path: Path, relative_path: str) -> List[CASMatch]:
    pages = _extract_pdf_pages(path)
    matches: List[CASMatch] = []
    for page_index, page_text in enumerate(pages, start=1):
        if not page_text.strip():
            continue
        matches.extend(
            _matches_from_text(
                page_text,
                source_file=path.name,
                relative_path=relative_path,
                file_type="pdf",
                location=f"page {page_index}",
            )
        )
    return matches


def _extract_pdf_pages(path: Path) -> List[str]:
    try:
        import pypdf

        reader = pypdf.PdfReader(str(path))
        return [page.extract_text() or "" for page in reader.pages]
    except ImportError:
        pass
    except Exception as exc:
        raise RuntimeError(f"pypdf could not read PDF: {exc}") from exc

    try:
        from pdfminer.high_level import extract_text

        text = extract_text(str(path))
        return [text or ""]
    except ImportError as exc:
        raise RuntimeError("Install pypdf or pdfminer.six to read PDF files") from exc
    except Exception as exc:
        raise RuntimeError(f"pdfminer could not read PDF: {exc}") from exc


def _extract_from_docx(path: Path, relative_path: str) -> List[CASMatch]:
    blocks = _read_docx_blocks(path)
    matches: List[CASMatch] = []
    for location, text in blocks:
        matches.extend(
            _matches_from_text(
                text,
                source_file=path.name,
                relative_path=relative_path,
                file_type="docx",
                location=location,
            )
        )
    return matches


def _read_docx_blocks(path: Path) -> List[Tuple[str, str]]:
    try:
        from docx import Document  # type: ignore

        document = Document(str(path))
        blocks: List[Tuple[str, str]] = []
        for index, paragraph in enumerate(document.paragraphs, start=1):
            text = (paragraph.text or "").strip()
            if text:
                blocks.append((f"paragraph {index}", text))
        if blocks:
            return blocks
    except ImportError:
        pass
    except Exception:
        pass

    blocks = []
    with zipfile.ZipFile(path) as archive:
        xml_names = [
            name
            for name in archive.namelist()
            if name.startswith("word/")
            and name.endswith(".xml")
            and Path(name).name in {"document.xml", "header1.xml", "header2.xml", "footer1.xml", "footer2.xml"}
        ]
        for xml_name in xml_names:
            root = ET.fromstring(archive.read(xml_name))
            paragraph_count = 0
            for paragraph in root.findall(".//w:p", _WORD_NS):
                parts = [node.text or "" for node in paragraph.findall(".//w:t", _WORD_NS)]
                text = "".join(parts).strip()
                if not text:
                    continue
                paragraph_count += 1
                label = Path(xml_name).stem
                blocks.append((f"{label} paragraph {paragraph_count}", text))
    return blocks


def _extract_from_xlsx(path: Path, relative_path: str) -> List[CASMatch]:
    try:
        return _extract_from_xlsx_openpyxl(path, relative_path)
    except ImportError:
        return _extract_from_xlsx_zip(path, relative_path)
    except Exception:
        return _extract_from_xlsx_zip(path, relative_path)


def _extract_from_xlsx_openpyxl(path: Path, relative_path: str) -> List[CASMatch]:
    from openpyxl import load_workbook  # type: ignore

    workbook = load_workbook(filename=str(path), read_only=True, data_only=True)
    matches: List[CASMatch] = []
    try:
        for sheet in workbook.worksheets:
            for row in sheet.iter_rows():
                for cell in row:
                    value = cell.value
                    if value in (None, ""):
                        continue
                    matches.extend(
                        _matches_from_text(
                            str(value),
                            source_file=path.name,
                            relative_path=relative_path,
                            file_type="xlsx",
                            location=f"sheet {sheet.title}!{cell.coordinate}",
                        )
                    )
    finally:
        workbook.close()
    return matches


def _extract_from_xlsx_zip(path: Path, relative_path: str) -> List[CASMatch]:
    with zipfile.ZipFile(path) as archive:
        workbook_root = ET.fromstring(archive.read("xl/workbook.xml"))
        rel_root = ET.fromstring(archive.read("xl/_rels/workbook.xml.rels"))
        rel_map = {
            rel.get("Id"): rel.get("Target", "")
            for rel in rel_root.findall(f"{{{_PKG_REL_NS}}}Relationship")
        }
        shared_strings = _read_xlsx_shared_strings(archive)
        matches: List[CASMatch] = []
        for sheet in workbook_root.findall(".//main:sheets/main:sheet", _XLSX_NS):
            sheet_name = sheet.get("name") or "Sheet"
            rel_id = sheet.get(f"{{{_XLSX_NS['rel']}}}id")
            target = rel_map.get(rel_id or "", "")
            if not target:
                continue
            normalized_target = target.lstrip("/")
            if not normalized_target.startswith("xl/"):
                normalized_target = f"xl/{normalized_target}"
            sheet_root = ET.fromstring(archive.read(normalized_target))
            for cell in sheet_root.findall(".//main:sheetData/main:row/main:c", _XLSX_NS):
                cell_ref = cell.get("r") or "unknown"
                value = _xlsx_cell_value(cell, shared_strings)
                if not value:
                    continue
                matches.extend(
                    _matches_from_text(
                        value,
                        source_file=path.name,
                        relative_path=relative_path,
                        file_type="xlsx",
                        location=f"sheet {sheet_name}!{cell_ref}",
                    )
                )
        return matches


def _read_xlsx_shared_strings(archive: zipfile.ZipFile) -> List[str]:
    if "xl/sharedStrings.xml" not in archive.namelist():
        return []
    root = ET.fromstring(archive.read("xl/sharedStrings.xml"))
    values: List[str] = []
    for item in root.findall(".//main:si", _XLSX_NS):
        parts = [node.text or "" for node in item.findall(".//main:t", _XLSX_NS)]
        values.append("".join(parts))
    return values


def _xlsx_cell_value(cell: ET.Element, shared_strings: Sequence[str]) -> str:
    cell_type = cell.get("t") or ""
    if cell_type == "inlineStr":
        return "".join(node.text or "" for node in cell.findall(".//main:is/main:t", _XLSX_NS)).strip()
    value = (cell.findtext("main:v", default="", namespaces=_XLSX_NS) or "").strip()
    if not value:
        return ""
    if cell_type == "s":
        try:
            return str(shared_strings[int(value)])
        except (IndexError, ValueError):
            return value
    return value


def _extract_from_xls(path: Path, relative_path: str) -> Tuple[List[CASMatch], List[str]]:
    try:
        import pandas as pd
    except ImportError as exc:
        return [], [f"Skipped {path.name}: install pandas with xlrd support to read .xls files ({exc})"]

    try:
        sheet_map = pd.read_excel(path, sheet_name=None, header=None, dtype=str)
    except Exception as exc:
        return [], [f"Skipped {path.name}: could not read .xls file ({exc})"]

    matches: List[CASMatch] = []
    for sheet_name, frame in sheet_map.items():
        for row_index, row in enumerate(frame.itertuples(index=False, name=None), start=1):
            for column_index, value in enumerate(row, start=1):
                if value in (None, ""):
                    continue
                matches.extend(
                    _matches_from_text(
                        str(value),
                        source_file=path.name,
                        relative_path=relative_path,
                        file_type="xls",
                        location=f"sheet {sheet_name}!R{row_index}C{column_index}",
                    )
                )
    return matches, []